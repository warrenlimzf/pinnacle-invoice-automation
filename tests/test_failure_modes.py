"""Failure-mode regression tests — synthetic PDFs only, no client data needed.

Run: python tests/test_failure_modes.py   (any machine, no samples/ required)

Covers the three ways a statement PDF used to vanish WITHOUT leaving a row in
the Excel (the "empty tab" bug the colleague hit on 2026-07-07):
  1. a password-protected PDF          -> must write a visible FAILED row
  2. the OCR engine crashing mid-read  -> must not kill the file, row written
  3. a scanned PDF with no OCR installed -> the flag must name that cause
"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import fitz
from openpyxl import load_workbook

import config
from shared.readers import pdf_reader

COL_SOURCE, COL_FLAGS = 9, 12


def _tmp_env() -> Path:
    """Point all outputs at a throwaway folder so tests never touch real files."""
    tmp = Path(tempfile.mkdtemp(prefix="pinnacle_test_"))
    config.MASTER_WORKBOOK = tmp / "nav_master.xlsx"
    config.SNAPSHOT_DIR = tmp / "snapshots"
    config.verification_docx = lambda bank: tmp / f"{bank}_verification.docx"
    return tmp


def _make_pdf(path: Path, text: str = "", encrypted: bool = False) -> None:
    doc = fitz.open()
    page = doc.new_page()
    if text:
        page.insert_text((72, 72), text)
    kwargs = {}
    if encrypted:
        kwargs = dict(encryption=fitz.PDF_ENCRYPT_AES_256,
                      user_pw="secret", owner_pw="secret")
    doc.save(str(path), **kwargs)
    doc.close()


def _flags_of(bank: str, source_name: str):
    wb = load_workbook(config.MASTER_WORKBOOK)
    ws = wb[bank]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=COL_SOURCE).value or "") == source_name:
            return str(ws.cell(row=r, column=COL_FLAGS).value or "")
    return None


def test_encrypted_pdf_writes_failed_row(tmp: Path) -> None:
    from shared.process import process_pdf
    pdf = tmp / "locked.pdf"
    _make_pdf(pdf, text="hello", encrypted=True)
    results = process_pdf("UBS", pdf)            # must NOT raise
    flags = _flags_of("UBS", "locked.pdf")
    assert flags is not None, "no row written for a password-protected PDF"
    assert "password" in flags.lower(), f"flag doesn't name the cause: {flags}"
    assert all(r.failed for r in results), \
        "failed result not marked failed — the runner would skip retrying it"
    print("PASS  encrypted PDF -> visible FAILED row naming the password")


def test_ocr_crash_is_contained(tmp: Path) -> None:
    from shared.process import process_pdf
    pdf = tmp / "scan_ocr_crash.pdf"
    _make_pdf(pdf)                               # page with no text layer

    class Boom:
        def __call__(self, *a, **k):
            raise RuntimeError("onnxruntime blew up")

    pdf_reader._ocr_engine, pdf_reader._ocr_import_failed = Boom(), False
    try:
        process_pdf("UBS", pdf)                  # must NOT raise
    finally:
        pdf_reader._ocr_engine, pdf_reader._ocr_import_failed = None, False
    flags = _flags_of("UBS", "scan_ocr_crash.pdf")
    assert flags is not None, "no row written when the OCR engine crashed"
    print("PASS  OCR crash -> contained, row still written")


def test_scanned_without_ocr_names_the_cause(tmp: Path) -> None:
    from shared.process import process_pdf
    pdf = tmp / "scan_no_ocr.pdf"
    _make_pdf(pdf)                               # page with no text layer
    pdf_reader._ocr_engine, pdf_reader._ocr_import_failed = None, True
    try:
        process_pdf("UBS", pdf)                  # must NOT raise
    finally:
        pdf_reader._ocr_import_failed = False
    flags = _flags_of("UBS", "scan_no_ocr.pdf") or ""
    assert "scan" in flags.lower() or "ocr" in flags.lower(), \
        f"flag doesn't tell her the PDF is a scan / OCR is missing: {flags}"
    print("PASS  scan without OCR -> flag names the cause")


def test_ubs_single_portfolio_statement(tmp: Path) -> None:
    """UBS also exports ONE PDF per portfolio: same asset-class table on the
    overview page but with NO 'Portfolio NN' heading (colleague's real files,
    2026-07-07). The parser must read that table instead of coming back empty."""
    from banks.UBS.parser import parse
    pdf = tmp / "ubs_single.pdf"
    doc = fitz.open()
    page = doc.new_page()
    rows = [
        "Total assets",
        "Portfolio number 546-123456-02",
        "Valued in USD",
        "Assets as of 28.02.2026",
        "Asset class Market value Accrued interest Total",
        "Liquidity 2 866 000 1 500 2 867 500",
        "Bonds 5 000 000 0 5 000 000",
        "Gross assets 14 500 000 1 500 14 501 500",
        "Liabilities -2 450 000 0 -2 450 000",
        "Net assets 12 050 000 1 500 12 051 500",
    ]
    for i, text in enumerate(rows):
        page.insert_text((72, 90 + i * 24), text, fontsize=10)
    doc.save(str(pdf))
    doc.close()

    res = parse(pdf)[0]
    assert res.account_no == "546-123456-02", res.account_no
    assert res.currency == "USD", res.currency
    assert res.gross_nav == 14_500_000, res.gross_nav
    assert res.net_nav == 12_050_000, res.net_nav
    assert res.liquidity == 2_866_000, res.liquidity
    assert res.liabilities == -2_450_000, res.liabilities
    print("PASS  UBS one-portfolio-per-PDF statement -> table read directly")


def test_ubs_two_tables_takes_first(tmp: Path) -> None:
    """When a heading-less UBS page shows SEVERAL asset tables, the client's own
    portfolio is printed first (per the suffix) — read the FIRST table only and
    never mix rows across tables."""
    from banks.UBS.parser import parse
    pdf = tmp / "ubs_two_tables.pdf"
    doc = fitz.open()
    page = doc.new_page()
    rows = [
        "Total assets",
        "Portfolio number 546-123456-03",
        "Valued in USD",
        "Liquidity 1 000 000 0 1 000 000",
        "Equities 9 000 000 0 9 000 000",
        "Gross assets 10 000 000 0 10 000 000",
        "Liabilities -1 000 000 0 -1 000 000",
        "Net assets 9 000 000 0 9 000 000",
        # a second portfolio's table further down, no heading either
        "Liquidity 77 0 77",
        "Gross assets 99 0 99",
        "Net assets 98 0 98",
    ]
    for i, text in enumerate(rows):
        page.insert_text((72, 90 + i * 24), text, fontsize=10)
    doc.save(str(pdf))
    doc.close()

    res = parse(pdf)[0]
    assert res.gross_nav == 10_000_000, res.gross_nav
    assert res.net_nav == 9_000_000, res.net_nav
    assert res.liquidity == 1_000_000, res.liquidity
    assert res.liabilities == -1_000_000, res.liabilities
    assert any("FIRST" in f for f in res.flags), res.flags
    print("PASS  UBS several heading-less tables -> first table only, flagged")


def main() -> None:
    failures = 0
    for test in (test_encrypted_pdf_writes_failed_row,
                 test_ocr_crash_is_contained,
                 test_scanned_without_ocr_names_the_cause,
                 test_ubs_single_portfolio_statement,
                 test_ubs_two_tables_takes_first):
        try:
            test(_tmp_env())
        except Exception as exc:
            failures += 1
            print(f"FAIL  {test.__name__}: {type(exc).__name__}: {exc}")
    if failures:
        sys.exit(f"{failures} failure-mode test(s) failed")
    print("All failure-mode tests passed.")


if __name__ == "__main__":
    main()
