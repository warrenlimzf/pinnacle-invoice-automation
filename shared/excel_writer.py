"""Owns the master workbook (output/nav_master.xlsx) — one tab per bank.

Finance convention (as requested):
  - HARDCODED values (read straight off the PDF) are BLUE
  - FORMULA cells (anything derived, e.g. LGT's reconstructed Gross NAV or a
    Gross-equals-Net case) are BLACK

Column order, ranked by importance (per the supervisor's guide):
  A Statement Date | B Account No | C Currency | D Gross NAV | E Net NAV |
  F Liquidity (additional -> last of the data block) |
  G Liabilities (BoS only) | H Check (formula: Gross + Liabilities - Net = 0) |
  I Source PDF | J Page | K Updated At | L Flags | N.. LGT add-back items

The block your colleague copies into her own master is A:F.
Each row = one statement PDF. Re-processing the same file UPDATES its row.
"""
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from shared.logging_setup import get_logger
from shared.model import ClientResult

log = get_logger("excel_writer")

# Column layout (1-based)
COL_DATE = 1      # A
COL_ACCOUNT = 2   # B
COL_CCY = 3       # C
COL_GROSS = 4     # D
COL_NET = 5       # E
COL_LIQ = 6       # F
COL_LIAB = 7      # G
COL_CHECK = 8     # H
COL_SOURCE = 9    # I
COL_PAGE = 10     # J
COL_UPDATED = 11  # K
COL_FLAGS = 12    # L
ADDBACK_START = 14        # N onwards (LGT add-back line items)
ADDBACK_CLEAR_TO = 29     # clear up to col AC when updating a row

HEADERS = ["Statement Date", "Account No", "Currency", "Gross NAV", "Net NAV",
           "Liquidity", "Liabilities", "Check (=0)", "Source PDF", "Page",
           "Updated At", "Flags"]

NUM_FMT = "#,##0.00"
BLUE = Font(color="FF00008B")      # dark blue = hardcoded inputs off the PDF
BLACK = Font(color="FF000000")     # black = formulas / derived
BOLD = Font(bold=True)


def _ensure_workbook(path: Path, banks: List[str]) -> Workbook:
    wb = load_workbook(path) if path.exists() else Workbook()
    if wb.sheetnames == ["Sheet"]:
        wb.remove(wb.active)
    for bank in banks:
        if bank not in wb.sheetnames:
            ws = wb.create_sheet(title=bank)
            _write_headers(ws, bank)
        else:
            ws = wb[bank]
            if str(ws.cell(row=1, column=1).value or "") != HEADERS[0]:
                log.warning(f"[{bank}] column layout has changed — rewriting the "
                            "header row. If old rows look misaligned, delete "
                            f"{path.name} and run run_once with --redo to rebuild.")
                _write_headers(ws, bank)
    return wb


def _write_headers(ws, bank: str) -> None:
    for c, head in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=head)
        cell.font = BOLD
    if bank == "LGT":
        note = ws.cell(row=1, column=ADDBACK_START,
                       value="LGT add-back line items (negative figures) — "
                             "Gross NAV = Net NAV with these added back →")
        note.font = BOLD


def _find_row(ws, source_name: str):
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=COL_SOURCE).value or "").strip() == source_name:
            return r
    return None


def _put(ws, r, col, value, font=BLUE, fmt=None):
    cell = ws.cell(row=r, column=col, value=value)
    cell.font = font
    if fmt:
        cell.number_format = fmt
    return cell


def write_result(path, bank: str, banks: List[str],
                 result: ClientResult, page) -> None:
    """Insert/update one statement's row. Raises PermissionError if the workbook
    is open in Excel (Windows locks it) — caller logs and retries later."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = _ensure_workbook(path, banks)
    ws = wb[bank]

    source_name = Path(result.source_pdf).name
    r = _find_row(ws, source_name) or (ws.max_row + 1)

    # --- the data block A:F (everything read off the PDF = blue) -------------
    if result.statement_date:
        _put(ws, r, COL_DATE, result.statement_date)
    if result.account_no:
        _put(ws, r, COL_ACCOUNT, result.account_no)
    if result.currency:
        _put(ws, r, COL_CCY, result.currency)
    if result.net_nav is not None:
        _put(ws, r, COL_NET, result.net_nav, fmt=NUM_FMT)
    if result.liquidity is not None:
        _put(ws, r, COL_LIQ, result.liquidity, fmt=NUM_FMT)

    # clear any stale add-back cells from a previous run of this row
    for col in range(ADDBACK_START, ADDBACK_CLEAR_TO + 1):
        cell = ws.cell(row=r, column=col)
        cell.value, cell.comment = None, None

    if result.gross_is_formula:
        # Derived Gross (LGT, or a UBS portfolio with no liabilities):
        # write each add-back (blue, labelled by a cell comment), then a live
        # formula Gross = Net - SUM(add-backs). Add-backs are negative numbers,
        # so subtracting them ADDS their absolute value back. No add-backs ->
        # Gross simply = Net.
        last_col = None
        for i, ab in enumerate(result.addbacks):
            col = ADDBACK_START + i
            cell = _put(ws, r, col, ab.value, fmt=NUM_FMT)
            cell.comment = Comment(f"{ab.label}", "automation")
            last_col = col
        net_ref = f"{get_column_letter(COL_NET)}{r}"
        if last_col:
            rng = (f"{get_column_letter(ADDBACK_START)}{r}:"
                   f"{get_column_letter(last_col)}{r}")
            formula = f"={net_ref}-SUM({rng})"
        else:
            formula = f"={net_ref}"   # no negatives -> Gross = Net
        _put(ws, r, COL_GROSS, formula, font=BLACK, fmt=NUM_FMT)
    elif result.gross_nav is not None:
        # UBS / BoS: Gross is read straight off the PDF -> hardcoded blue.
        _put(ws, r, COL_GROSS, result.gross_nav, fmt=NUM_FMT)

    # --- BoS audit check: Gross + Liabilities - Net should be 0 --------------
    if result.liabilities is not None:
        _put(ws, r, COL_LIAB, result.liabilities, fmt=NUM_FMT)
        g = get_column_letter(COL_GROSS)
        l = get_column_letter(COL_LIAB)
        n = get_column_letter(COL_NET)
        cell = _put(ws, r, COL_CHECK, f"=ROUND({g}{r}+{l}{r}-{n}{r},2)",
                    font=BLACK, fmt=NUM_FMT)
        cell.comment = Comment("Investment Assets + Liabilities - Net NAV. "
                               "Should be 0.00.", "automation")

    ws.cell(row=r, column=COL_SOURCE, value=source_name)
    ws.cell(row=r, column=COL_PAGE, value=page)
    ws.cell(row=r, column=COL_UPDATED,
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=r, column=COL_FLAGS, value="; ".join(result.flags))

    wb.save(path)
